#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""battery_catalog.html 의 시드 데이터 → 엑셀 워크북.

앱에서 보던 것(카달로그·규격·제조사·차량적합성·지역·크로스레퍼런스)을
엑셀로 옮긴다.  파생지표와 집계는 값이 아니라 **수식**으로 넣어서,
숫자를 고치면 갭·점수·커버리지가 따라 움직인다.

사용법:
    node tools/dump_seed.js > seed.json      # 또는 아래 --seed 로 경로 지정
    python3 tools/build_workbook.py seed.json 배터리_카달로그.xlsx

주의: 사내 데이터(고객품번·납품처·오더량)는 여기 들어가지 않는다.
      그건 앱의 CSV 임포트로만 다루고, 이 저장소·이 파일에는 남기지 않는다.
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='14263D')
HDR_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=10)
INPUT_FILL = PatternFill('solid', fgColor='FFF9DB')   # 사람이 채우는 칸
OWN_FILL = PatternFill('solid', fgColor='FDE9EC')     # 자사
TITLE = Font(name=FONT, bold=True, size=14)
SUB = Font(name=FONT, color='667487', size=9)
BOLD = Font(name=FONT, bold=True, size=10)
BODY = Font(name=FONT, size=10)
THIN = Side(style='thin', color='DFE5EC')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def head(ws, cols, row=1):
    """컬럼 정의 [(제목, 너비, 숫자서식|None, 입력칸여부)] 를 헤더로 쓴다."""
    for i, (title, width, _fmt, is_input) in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=title)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = width
        if is_input:
            ws.cell(row=row, column=i).comment = None
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def write_rows(ws, cols, rows, start=2):
    for r, values in enumerate(rows, start):
        for i, (v, (_t, _w, fmt, is_input)) in enumerate(zip(values, cols), 1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = BODY
            c.border = BOX
            if fmt:
                c.number_format = fmt
            if is_input and (v is None or v == ''):
                c.fill = INPUT_FILL


def build(seed, out):
    S = seed
    models = S['MODELS']
    specs = S['SPECS']
    makers = S['MAKERS']
    fits = S['FITMENT']
    regions = S['REGIONS']
    techs = S['TECHS']
    rgn_name = {r['k']: r['n'] for r in regions}
    mk_name = {m['k']: m['n'] for m in makers}
    spec_of = {s['k']: s for s in specs}
    HOME = 'sebang'

    n = len(models)
    LAST = n + 1                      # 카달로그 데이터 마지막 행
    CAT = '카달로그'
    rng = lambda col: "%s!$%s$2:$%s$%d" % (CAT, col, col, LAST)

    wb = Workbook()

    # ── 읽어보기 ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = '읽어보기'
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 96
    ws['A1'] = '글로벌 SLI 배터리 카달로그'
    ws['A1'].font = TITLE
    ws['A2'] = 'AGM · EFB · 일반액식(FLD) 시동용 배터리. battery_catalog.html 과 같은 데이터.'
    ws['A2'].font = SUB
    lines = [
        ('', ''),
        ('시트 구성', ''),
        ('카달로그', '제조사·유통사별 SLI 모델 %d종. 왼쪽부터 제품 → 사이즈(규격·치수) → 용량·CCA 순으로 놓았다. '
                  '판매지역은 지역마다 열이 따로 있어서(O 표시) 필터로 바로 뽑을 수 있다.' % n),
        ('지역별판매', '모델 × 판매지역을 한 행씩 편 표. "유럽에서 파는 LN3, 70Ah 이상" 같은 걸 필터 두 번으로 뽑는다. 피벗테이블 원본으로 쓰면 된다.'),
        ('사이즈별요약', '규격마다 치수·용량범위·CCA범위·지역별 취급 수를 한 줄로. 어떤 사이즈가 어디서 얼마짜리로 팔리는지 보는 표.'),
        ('규격마스터', '규격(사이즈)별 표준 치수 %d종. 카달로그 시트의 치수는 여기서 끌어온다.' % len(specs)),
        ('브랜드마켓', '같은 브랜드가 나라마다 다른 시리즈를 판다(Bosch 유럽 S3~S6 vs 북미 BCI vs 인도 S4+/T4/M6). '
                   '그 단위로 실제 카탈로그 주소와 수집상태를 관리한다.'),
        ('제조사', '제조사·유통사 %d곳과 카달로그 수집 현황.' % len(makers)),
        ('차량적합성', '차종이 요구하는 규격·용량·CCA·기술과, 그 요구를 만족하는 모델 수.'),
        ('지역커버리지', '판매지역 × 기술 매트릭스 + 지역별 취급 제조사·커버 규격수. 라인업 공백을 보는 표.'),
        ('크로스레퍼런스', '같은 규격에서 자사 최고 스펙과 타사 최고 스펙의 격차.'),
        ('설정', 'SAE→EN 환산계수. 이 값을 바꾸면 카달로그의 CCA 환산이 전부 다시 계산된다.'),
        ('', ''),
        ('근거 등급 (카달로그 Z열)', ''),
        ('공개스펙', '제조사가 공개한 스펙. 그대로 써도 되는 값.'),
        ('규격표준', '규격(BCI/DIN/JIS) 표준값.'),
        ('추정', '공개 자료로 채운 추정값. 원본 카달로그 대조 전에는 영업·기술 자료로 쓰지 말 것.'),
        ('라인업슬롯', '브랜드·시리즈·규격 커버리지만 확인한 칸. 실제 모델번호와 스펙은 카달로그를 받아야 채워진다.'),
        ('', ''),
        ('채워 넣는 칸 (연노랑)', ''),
        ('카달로그 근거문서·근거페이지', '어느 카달로그 몇 판 몇 쪽에서 본 값인지 적는다(맨 오른쪽 두 열).'),
        ('제조사 J~P열', '카달로그명 / 판 / 발행연도 / PDF파일명 / SHA-256 / 페이지수 / 수집일.'),
        ('', '예시 →  Exide Technical Guide | Edition 5 | 2019 | exide_guide_ed5.pdf | 7e76bbee70d6… | 84 | 2026-07-28'),
        ('', ''),
        ('주의', ''),
        ('CCA 환산', 'EN(10초/7.5V)과 SAE(30초/7.2V)는 시험법이 달라 정확한 환산식이 없다. R·S열의 환산값은 비교용 근사이고, 고객 제시에는 P·Q열의 원 표기값과 표기 규격을 써야 한다.'),
        ('사이클 수명', '제조사마다 DoD 기준이 달라 숫자를 직접 비교하면 안 된다.'),
        ('차량 적합성', '같은 차종도 연식·트림·사양에 따라 요구가 다르다. 최종 적용은 OEM 매뉴얼로 확인할 것.'),
        ('사내 데이터', '고객품번·납품처·오더량은 이 파일에 없다. 그건 앱의 CSV 임포트로만 다룬다.'),
        ('수식으로 만든 파일', '파생지표와 집계는 값이 아니라 수식이다. 엑셀·구글시트에서 열면 자동으로 계산된다. '
                          '계산값이 파일에 저장돼 있지는 않아서, 계산을 안 하는 미리보기 도구에서는 그 칸이 비어 보일 수 있다.'),
        ('다시 만들기', 'node tools/dump_seed.js battery_catalog.html seed.json  →  '
                    'python3 tools/build_workbook.py seed.json 배터리_카달로그.xlsx'),
    ]
    r = 3
    for a, b in lines:
        ws.cell(row=r, column=1, value=a).font = BOLD if b == '' and a else BODY
        ws.cell(row=r, column=2, value=b).font = BODY
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        r += 1

    # ── 설정 ──────────────────────────────────────────────────────────────
    st = wb.create_sheet('설정')
    st.column_dimensions['A'].width = 26
    st.column_dimensions['B'].width = 12
    st.column_dimensions['C'].width = 70
    for i, t in enumerate(['항목', '값', '설명'], 1):
        c = st.cell(row=1, column=i, value=t); c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BOX
    st['A2'] = 'SAE→EN 환산계수'; st['B2'] = 0.95
    st['C2'] = 'SAE 표기 CCA 에 곱해 EN 으로 본다. 시험법이 달라 정확한 식이 없는 근사값이다(기본 0.95).'
    st['A3'] = '자사 제조사'; st['B3'] = mk_name.get(HOME, HOME)
    st['C3'] = '카달로그 A열의 자사/타사 구분과 크로스레퍼런스의 갭 계산 기준.'
    for row in st.iter_rows(min_row=2, max_row=3, max_col=3):
        for c in row:
            c.font = BODY; c.border = BOX
    st['B2'].fill = INPUT_FILL; st['B2'].number_format = '0.00'
    st['B3'].fill = INPUT_FILL

    # ── 규격마스터 ────────────────────────────────────────────────────────
    sp_cols = [('규격그룹', 12, None, False), ('규격체계', 12, None, False),
               ('길이(mm)', 10, '#,##0', False), ('폭(mm)', 10, '#,##0', False),
               ('높이(mm)', 10, '#,##0', False), ('부피(L)', 10, '0.0', False),
               ('통용 용량대(Ah)', 14, None, False), ('비고', 34, None, False)]
    sw = wb.create_sheet('규격마스터')
    head(sw, sp_cols)
    for i, s in enumerate(specs, 2):
        write_rows(sw, sp_cols, [[s['k'], s['sys'], s['L'], s['W'], s['H'], None, s.get('ah', ''), s.get('note', '')]], i)
        sw.cell(row=i, column=6, value='=C%d*D%d*E%d/1000000' % (i, i, i)).number_format = '0.0'
        sw.cell(row=i, column=6).font = BODY; sw.cell(row=i, column=6).border = BOX
    sw.auto_filter.ref = 'A1:H%d' % (len(specs) + 1)

    # ── 카달로그 ──────────────────────────────────────────────────────────
    # 보는 순서대로 놓는다: 누가 만든 무슨 제품인지 → 어떤 사이즈인지 →
    # 용량·CCA 가 얼마인지 → 어느 지역에서 파는지.  지역은 텍스트 한 칸에
    # 뭉치면 필터가 안 되므로 지역마다 열을 준다.
    cat_cols = [
        ('구분', 7, None, False), ('제조사', 26, None, False), ('브랜드', 13, None, False),
        ('모델', 30, None, False), ('기술', 9, None, False), ('용도', 11, None, False),
        ('규격그룹', 11, None, False), ('규격코드', 14, None, False), ('규격체계', 10, None, False),
        ('길이(mm)', 9, '#,##0', False), ('폭(mm)', 9, '#,##0', False), ('높이(mm)', 9, '#,##0', False),
        ('용량 C20(Ah)', 12, '#,##0', False), ('RC(분)', 8, '#,##0', False),
        ('CCA 표기(A)', 11, '#,##0', False), ('표기규격', 9, None, False),
        ('CCA(EN)', 9, '#,##0', False), ('CCA(SAE)', 9, '#,##0', False),
        ('무게(kg)', 9, '0.0', False), ('사이클', 8, '#,##0', False), ('전압', 6, '0', False),
        ('CCA/Ah', 9, '0.00', False), ('Wh/kg', 8, '0.0', False), ('Wh/L', 8, '0.0', False),
    ]
    RG0 = len(cat_cols) + 1                       # 지역 열 시작 (1-base)
    cat_cols += [(r['n'], 9, None, False) for r in regions]
    cat_cols += [('판매지역 수', 10, '#,##0', False), ('판매지역', 30, None, False),
                 ('근거등급', 11, None, False), ('근거문서', 24, None, True),
                 ('근거페이지', 10, None, True), ('비고', 20, None, False),
                 ('브랜드마켓', 14, None, False)]
    RGN_END = RG0 + len(regions) - 1
    cw = wb.create_sheet(CAT)
    head(cw, cat_cols)
    VER = {'pub': '공개스펙', 'std': '규격표준', 'est': '추정', 'slot': '라인업슬롯',
           'own': '사내설계치', 'user': '직접입력'}
    for i, m in enumerate(models, 2):
        own = '자사' if m['mk'] == HOME else '타사'
        mrg = set(m.get('rg', []))
        row = [own, mk_name.get(m['mk'], m['mk']), m.get('br', ''), m['mdl'],
               m.get('tech', ''), m.get('app', ''), m.get('grp', ''), m.get('sz', ''),
               None, None, None, None,
               m.get('ah'), m.get('rc'), m.get('cca'), m.get('cs', ''), None, None,
               m.get('wt'), m.get('cyc'), m.get('v', 12), None, None, None]
        row += ['O' if r['k'] in mrg else '' for r in regions]
        row += [None, ', '.join(r['n'] for r in regions if r['k'] in mrg),
                VER.get(m.get('ver'), m.get('ver', '')), '', '', m.get('note', ''),
                m.get('bm', '')]
        write_rows(cw, cat_cols, [row], i)
        # 규격마스터에서 끌어오는 값 (I 규격체계, J·K·L 치수)
        for col, src in (('I', 'B'), ('J', 'C'), ('K', 'D'), ('L', 'E')):
            cw['%s%d' % (col, i)] = ('=IFERROR(INDEX(규격마스터!$%s$2:$%s$%d,'
                                     'MATCH($G%d,규격마스터!$A$2:$A$%d,0)),"")'
                                     % (src, src, len(specs) + 1, i, len(specs) + 1))
        cw['Q%d' % i] = '=IF($P%d="EN",$O%d,ROUND($O%d*설정!$B$2,0))' % (i, i, i)
        cw['R%d' % i] = '=IF($P%d="SAE",$O%d,ROUND($O%d/설정!$B$2,0))' % (i, i, i)
        cw['V%d' % i] = '=IFERROR($Q%d/$M%d,"")' % (i, i)
        cw['W%d' % i] = '=IFERROR($U%d*$M%d/$S%d,"")' % (i, i, i)
        cw['X%d' % i] = '=IFERROR($U%d*$M%d*1000000/($J%d*$K%d*$L%d),"")' % (i, i, i, i, i)
        cw.cell(row=i, column=RGN_END + 1).value = (
            '=COUNTIF(%s%d:%s%d,"O")' % (get_column_letter(RG0), i, get_column_letter(RGN_END), i))
        for col, fmt in (('I', None), ('J', '#,##0'), ('K', '#,##0'), ('L', '#,##0'),
                         ('Q', '#,##0'), ('R', '#,##0'), ('V', '0.00'), ('W', '0.0'), ('X', '0.0'),
                         (get_column_letter(RGN_END + 1), '#,##0')):
            c = cw['%s%d' % (col, i)]
            c.font = BODY; c.border = BOX
            if fmt:
                c.number_format = fmt
        for j in range(RG0, RGN_END + 1):
            cw.cell(row=i, column=j).alignment = Alignment(horizontal='center')
        if own == '자사':
            cw['A%d' % i].fill = OWN_FILL
    cw.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(cat_cols)), LAST)
    # 사이즈·용량·CCA 가 한눈에 들어오게 모델명까지 고정
    cw.freeze_panes = 'E2'

    # ── 제조사 ────────────────────────────────────────────────────────────
    mk_cols = [
        ('제조사키', 12, None, False), ('제조사 / 유통사', 34, None, False), ('국가', 14, None, False),
        ('구분', 8, None, False), ('브랜드', 40, None, False), ('주요 판매지역', 26, None, False),
        ('사이트', 42, None, False), ('등록 모델수', 11, '#,##0', False), ('수집상태', 10, None, True),
        ('카달로그명', 26, None, True), ('판(edition)', 12, None, True), ('발행연도', 10, None, True),
        ('PDF 파일명', 26, None, True), ('SHA-256', 26, None, True), ('페이지수', 9, '#,##0', True),
        ('수집일', 12, None, True),
    ]
    mw = wb.create_sheet('제조사')
    head(mw, mk_cols)
    for i, m in enumerate(makers, 2):
        write_rows(mw, mk_cols, [[
            m['k'], m['n'], m.get('c', ''), '유통사' if m.get('t') == 'D' else '제조사',
            m.get('br', ''), ', '.join(rgn_name.get(x, x) for x in m.get('rg', [])),
            m.get('site', ''), None, m.get('st', ''), m.get('cat', ''), m.get('ed', ''),
            m.get('pub', ''), m.get('file', ''), m.get('sha', ''), m.get('pg', ''), m.get('dt', '')
        ]], i)
        c = mw['H%d' % i]
        c.value = '=COUNTIF(%s,$B%d)' % (rng('B'), i)
        c.font = BODY; c.border = BOX; c.number_format = '#,##0'
        if m['k'] == HOME:
            mw['B%d' % i].fill = OWN_FILL
    mw.auto_filter.ref = 'A1:P%d' % (len(makers) + 1)

    # ── 브랜드마켓 ────────────────────────────────────────────────────────
    # 같은 브랜드라도 나라마다 파는 시리즈와 규격체계가 다르다.
    # 카탈로그 주소도 그 단위로 따로 있어서 여기에 모아 둔다.
    bms = S.get('BRAND_MARKETS', [])
    if bms:
        bm_cols = [
            ('브랜드마켓키', 14, None, False), ('브랜드', 14, None, False),
            ('제조사', 24, None, False), ('지역', 12, None, False), ('마켓/국가', 16, None, False),
            ('규격체계', 16, None, False), ('판매 시리즈', 46, None, False),
            ('등록 모델', 10, '#,##0', False), ('소스 URL 수', 11, '#,##0', False),
            ('카탈로그 주소', 60, None, True), ('주소 확인', 11, None, False),
            ('피트먼트 검색', 40, None, True), ('수집상태', 11, None, True),
        ]
        bw = wb.create_sheet('브랜드마켓')
        head(bw, bm_cols)
        for i, b in enumerate(bms, 2):
            write_rows(bw, bm_cols, [[
                b['k'], b.get('br', ''), mk_name.get(b.get('mk'), b.get('mk', '')),
                rgn_name.get(b.get('rg'), b.get('rg', '')), b.get('ctry', ''),
                b.get('std', ''), b.get('series', ''), None, b.get('nsrc', ''),
                b.get('cat', ''), {'검색확인': '확인됨'}.get(b.get('catChk', ''), '경로추정' if b.get('cat') else ''),
                b.get('fit', ''), b.get('st', '미수집')
            ]], i)
            c = bw.cell(row=i, column=8, value='=COUNTIF(%s,$A%d)' % (rng('AP'), i))
            c.font = BODY; c.border = BOX; c.number_format = '#,##0'
        bw.auto_filter.ref = 'A1:M%d' % (len(bms) + 1)

    # ── 차량적합성 ────────────────────────────────────────────────────────
    ft_cols = [
        ('시장', 8, None, False), ('차량 브랜드', 16, None, False), ('차종', 22, None, False),
        ('연식', 13, None, False), ('엔진/사양', 22, None, False), ('ISG', 7, None, False),
        ('요구 규격', 11, None, False), ('요구 Ah', 9, '#,##0', False),
        ('요구 CCA(EN)', 12, '#,##0', False), ('요구 기술', 10, None, False),
        ('대체 / 비고', 30, None, False), ('확신도', 10, None, False),
        ('적합 모델수', 11, '#,##0', False), ('자사 대응', 10, '#,##0', False),
    ]
    fw = wb.create_sheet('차량적합성')
    head(fw, ft_cols)
    for i, f in enumerate(fits, 2):
        write_rows(fw, ft_cols, [[
            f.get('mkt', ''), f.get('bd', ''), f.get('mdl', ''), f.get('yr', ''), f.get('eng', ''),
            '있음' if f.get('isg') else '없음', f.get('grp', ''), f.get('ah'), f.get('cca'),
            f.get('tech', ''), f.get('alt', ''),
            '업계 통용' if f.get('conf') == 'H' else '대조 필요', None, None
        ]], i)
        tech = f.get('tech', '')
        if tech == 'AGM':
            tcond = '*ISNUMBER(SEARCH("AGM",%s))' % rng('E')
        elif tech == 'EFB':
            tcond = '*((ISNUMBER(SEARCH("AGM",%s))+ISNUMBER(SEARCH("EFB",%s)))>0)' % (rng('E'), rng('E'))
        else:
            tcond = ''
        base = ('(%s=$G%d)*(%s>=$H%d*0.95)*(%s>=$I%d*0.95)%s'
                % (rng('G'), i, rng('M'), i, rng('Q'), i, tcond))
        fw['M%d' % i] = '=SUMPRODUCT(%s)' % base
        fw['N%d' % i] = '=SUMPRODUCT(%s*(%s="자사"))' % (base, rng('A'))
        for col in ('M', 'N'):
            c = fw['%s%d' % (col, i)]
            c.font = BODY; c.border = BOX; c.number_format = '#,##0'
    fw.auto_filter.ref = 'A1:N%d' % (len(fits) + 1)

    # ── 지역커버리지 ──────────────────────────────────────────────────────
    # 카달로그의 지역 열(O 표시)을 그대로 센다. 텍스트를 뒤지지 않아 정확하다.
    rg_cols = ([('판매지역', 16, None, False)]
               + [(t, 11, '#,##0', False) for t in techs]
               + [('합계', 10, '#,##0', False), ('자사', 9, '#,##0', False),
                  ('타사', 9, '#,##0', False), ('자사 비중', 10, '0.0%', False),
                  ('취급 제조사', 11, '#,##0', False), ('커버 규격수', 11, '#,##0', False)])
    rw = wb.create_sheet('지역커버리지')
    head(rw, rg_cols)
    MEND = len(makers) + 1
    for i, rg in enumerate(regions, 2):
        col = get_column_letter(RG0 + i - 2)
        rgcol = '%s!$%s$2:$%s$%d' % (CAT, col, col, LAST)
        rw.cell(row=i, column=1, value=rg['n']).font = BODY
        rw.cell(row=i, column=1).border = BOX
        cells = []
        for j, t in enumerate(techs, 2):
            cells.append((j, '=COUNTIFS(%s,"O",%s,"%s")' % (rgcol, rng('E'), t), '#,##0'))
        last_t = get_column_letter(1 + len(techs))
        base = 2 + len(techs)
        cells += [
            (base,     '=SUM(B%d:%s%d)' % (i, last_t, i), '#,##0'),
            (base + 1, '=COUNTIFS(%s,"O",%s,"자사")' % (rgcol, rng('A')), '#,##0'),
            (base + 2, '=COUNTIFS(%s,"O",%s,"타사")' % (rgcol, rng('A')), '#,##0'),
            (base + 3, '=IFERROR(%s%d/%s%d,"")' % (get_column_letter(base + 1), i,
                                                   get_column_letter(base), i), '0.0%'),
            (base + 4, '=SUMPRODUCT(--(COUNTIFS(%s,"O",%s,제조사!$B$2:$B$%d)>0))'
                       % (rgcol, rng('B'), MEND), '#,##0'),
            (base + 5, '=SUMPRODUCT(--(COUNTIFS(%s,"O",%s,규격마스터!$A$2:$A$%d)>0))'
                       % (rgcol, rng('G'), len(specs) + 1), '#,##0'),
        ]
        for cidx, formula, fmt in cells:
            c = rw.cell(row=i, column=cidx, value=formula)
            c.font = BODY; c.border = BOX; c.number_format = fmt

    # ── 지역별 판매 (롱포맷) ──────────────────────────────────────────────
    # 모델 × 판매지역을 한 행씩 편다.  "유럽에서 파는 LN3, 70Ah 이상" 같은 걸
    # 필터 두 번으로 뽑을 수 있고, 피벗테이블 원본으로도 이 시트를 쓴다.
    lg_cols = [
        ('판매지역', 14, None, False), ('구분', 7, None, False), ('제조사', 26, None, False),
        ('브랜드', 13, None, False), ('모델', 30, None, False), ('기술', 9, None, False),
        ('용도', 11, None, False), ('규격그룹', 11, None, False), ('규격코드', 14, None, False),
        ('치수 L×W×H', 17, None, False), ('용량 C20(Ah)', 12, '#,##0', False),
        ('RC(분)', 8, '#,##0', False), ('CCA 표기(A)', 11, '#,##0', False),
        ('표기규격', 9, None, False), ('CCA(EN)', 9, '#,##0', False),
        ('무게(kg)', 9, '0.0', False), ('근거등급', 11, None, False),
    ]
    lw = wb.create_sheet('지역별판매')
    head(lw, lg_cols)
    order = {r['k']: n for n, r in enumerate(regions)}
    long_rows = []
    for m in models:
        sp = spec_of.get(m.get('grp'), {})
        dim = ('%d×%d×%d' % (sp['L'], sp['W'], sp['H'])) if sp else ''
        for rk in m.get('rg', []):
            long_rows.append((order.get(rk, 99), m.get('grp') or '', -(m.get('cca') or 0), [
                rgn_name.get(rk, rk), '자사' if m['mk'] == HOME else '타사',
                mk_name.get(m['mk'], m['mk']), m.get('br', ''), m['mdl'],
                m.get('tech', ''), m.get('app', ''), m.get('grp', ''), m.get('sz', ''),
                dim, m.get('ah'), m.get('rc'), m.get('cca'), m.get('cs', ''), None,
                m.get('wt'), VER.get(m.get('ver'), m.get('ver', ''))]))
    long_rows.sort(key=lambda x: (x[0], x[1], x[2]))
    for i, (_a, _b, _c, vals) in enumerate(long_rows, 2):
        write_rows(lw, lg_cols, [vals], i)
        c = lw.cell(row=i, column=15, value='=IF($N%d="EN",$M%d,ROUND($M%d*설정!$B$2,0))' % (i, i, i))
        c.font = BODY; c.border = BOX; c.number_format = '#,##0'
        if vals[1] == '자사':
            lw.cell(row=i, column=2).fill = OWN_FILL
    lw.auto_filter.ref = 'A1:Q%d' % (len(long_rows) + 1)
    lw.freeze_panes = 'F2'
    LONG_END = len(long_rows) + 1

    # ── 사이즈별 요약 ─────────────────────────────────────────────────────
    # "이 사이즈는 몇 mm 이고, 용량·CCA 가 어디부터 어디까지이고,
    #  어느 지역에서 몇 개 팔리는지" 를 한 줄로 본다.
    sz_cols = ([('규격그룹', 12, None, False), ('규격체계', 11, None, False),
                ('길이(mm)', 9, '#,##0', False), ('폭(mm)', 9, '#,##0', False),
                ('높이(mm)', 9, '#,##0', False), ('부피(L)', 9, '0.0', False),
                ('통용 용량대(Ah)', 14, None, False), ('모델수', 8, '#,##0', False),
                ('용량 최소', 10, '#,##0', False), ('용량 최대', 10, '#,##0', False),
                ('CCA(EN) 최소', 12, '#,##0', False), ('CCA(EN) 최대', 12, '#,##0', False),
                ('자사 보유', 9, '#,##0', False)]
               + [(r['n'], 9, '#,##0', False) for r in regions])
    zw = wb.create_sheet('사이즈별요약')
    head(zw, sz_cols)
    used_all = [s2 for s2 in specs if any(m.get('grp') == s2['k'] for m in models)]
    for i, s2 in enumerate(used_all, 2):
        write_rows(zw, sz_cols, [[s2['k'], s2['sys'], s2['L'], s2['W'], s2['H'], None,
                                  s2.get('ah', '')] + [None] * (6 + len(regions))], i)
        f = {
            6: '=C%d*D%d*E%d/1000000' % (i, i, i),
            8: '=COUNTIF(%s,$A%d)' % (rng('G'), i),
            9: '=IF($H%d=0,"",100000-SUMPRODUCT(MAX((%s=$A%d)*(100000-%s))))'
               % (i, rng('G'), i, rng('M')),
            10: '=IF($H%d=0,"",SUMPRODUCT(MAX((%s=$A%d)*%s)))' % (i, rng('G'), i, rng('M')),
            11: '=IF($H%d=0,"",100000-SUMPRODUCT(MAX((%s=$A%d)*(100000-%s))))'
                % (i, rng('G'), i, rng('Q')),
            12: '=IF($H%d=0,"",SUMPRODUCT(MAX((%s=$A%d)*%s)))' % (i, rng('G'), i, rng('Q')),
            13: '=COUNTIFS(%s,$A%d,%s,"자사")' % (rng('G'), i, rng('A')),
        }
        for j, _r in enumerate(regions):
            col = get_column_letter(RG0 + j)
            f[14 + j] = ('=COUNTIFS(%s,$A%d,%s!$%s$2:$%s$%d,"O")'
                         % (rng('G'), i, CAT, col, col, LAST))
        for cidx, formula in f.items():
            c = zw.cell(row=i, column=cidx, value=formula)
            c.font = BODY; c.border = BOX
            c.number_format = sz_cols[cidx - 1][2] or 'General'
    zw.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(sz_cols)), len(used_all) + 1)

    # ── 크로스레퍼런스 ────────────────────────────────────────────────────
    xr_cols = [
        ('규격그룹', 12, None, False), ('규격체계', 11, None, False), ('치수 L×W×H', 16, None, False),
        ('모델수', 9, '#,##0', False), ('제조사수', 9, '#,##0', False),
        ('자사 최고 CCA', 12, '#,##0', False), ('타사 최고 CCA', 12, '#,##0', False),
        ('CCA 갭', 10, '0.0%', False),
        ('자사 최고 Ah', 12, '#,##0', False), ('타사 최고 Ah', 12, '#,##0', False),
        ('Ah 갭', 10, '0.0%', False),
        ('자사 AGM', 10, '#,##0', False), ('전체 AGM', 10, '#,##0', False),
    ]
    xw = wb.create_sheet('크로스레퍼런스')
    head(xw, xr_cols)
    used = [s for s in specs if any(m.get('grp') == s['k'] for m in models)]
    MEND = len(makers) + 1
    for i, s in enumerate(used, 2):
        sp = spec_of[s['k']]
        write_rows(xw, xr_cols, [[
            s['k'], s['sys'], '%d×%d×%d' % (sp['L'], sp['W'], sp['H']),
            None, None, None, None, None, None, None, None, None, None
        ]], i)
        f = {}
        f[4] = '=COUNTIF(%s,$A%d)' % (rng('G'), i)
        f[5] = ('=SUMPRODUCT(--(COUNTIFS(%s,$A%d,%s,제조사!$B$2:$B$%d)>0))'
                % (rng('G'), i, rng('B'), MEND))
        for col, cat_col, side in ((6, 'Q', '자사'), (7, 'Q', '타사'), (9, 'M', '자사'), (10, 'M', '타사')):
            # MAXIFS 는 엑셀 2016 이상 전용이라 구버전에서 #NAME? 가 된다.
            # SUMPRODUCT(MAX(...)) 는 2007 부터 배열 입력 없이도 도는 고전 패턴이다.
            f[col] = ('=IF(COUNTIFS(%s,$A%d,%s,"%s")=0,"",'
                      'SUMPRODUCT(MAX((%s=$A%d)*(%s="%s")*%s)))'
                      % (rng('G'), i, rng('A'), side,
                         rng('G'), i, rng('A'), side, rng(cat_col)))
        f[8] = '=IFERROR(($F%d-$G%d)/$G%d,"")' % (i, i, i)
        f[11] = '=IFERROR(($I%d-$J%d)/$J%d,"")' % (i, i, i)
        f[12] = ('=SUMPRODUCT((%s=$A%d)*(%s="자사")*ISNUMBER(SEARCH("AGM",%s)))'
                 % (rng('G'), i, rng('A'), rng('E')))
        f[13] = '=SUMPRODUCT((%s=$A%d)*ISNUMBER(SEARCH("AGM",%s)))' % (rng('G'), i, rng('E'))
        for col, formula in f.items():
            c = xw.cell(row=i, column=col, value=formula)
            c.font = BODY; c.border = BOX
            c.number_format = xr_cols[col - 1][2] or 'General'
    xw.auto_filter.ref = 'A1:M%d' % (len(used) + 1)

    # openpyxl 은 수식의 계산값을 못 넣는다. 열 때 전부 계산하도록 표시해 둔다.
    wb.calculation.fullCalcOnLoad = True
    wb.save(out)
    return {'models': n, 'specs': len(specs), 'makers': len(makers),
            'fitment': len(fits), 'xref': len(used)}


if __name__ == '__main__':
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    with open(sys.argv[1], encoding='utf-8') as fp:
        seed = json.load(fp)
    print(build(seed, sys.argv[2]), '→', sys.argv[2])
