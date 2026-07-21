#!/usr/bin/env python3
"""
BTS-600 현황 표 스크린샷 → 엑셀 (프로그램 무접촉, 안전)

용도(정정): '매일 값 기록'이 아니라 **시험 완료 감지**용.
  진행 중 값은 계속 변해 의미 없음. 이 표에서 필요한 건 Status가
  실행(CHA/DCH/PAU) → STO 로 바뀌는 '완료 전환'을 잡는 것.
  스크린샷은 프로그램 내부를 안 건드려 죽을 위험이 없으므로,
  주기적으로 화면만 찍어 '어느 회로가 방금 끝났는지'를 안전하게 감지한다.
  완료가 감지되면 → 그 회로 1개만 export → 관리대장·성적서 자동 처리.

demo: 실제 화면에서 비전으로 읽은 행들(상태 포함)을 엑셀에 기입.
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY = "004094"
HEAD = PatternFill("solid", fgColor=NAVY)
HEADF = Font(color="FFFFFF", bold=True, size=10)
RUN = PatternFill("solid", fgColor="FBF1DD")
NODATA = PatternFill("solid", fgColor="FDECEA")
THIN = Side(style="thin", color="D8DEE6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = ["회로", "시료", "프로그램", "상태", "사이클", "경과시간",
        "전압(V)", "전류(A)", "누적(Ah)", "방전용량(Ah)"]

# 물리적 불가값(센서/데이터 없음) 판별
def nodata(v):
    try:
        return float(v) < -200
    except (ValueError, TypeError):
        return False


def build(rows, out="prototype/out/grid_status.xlsx", when="화면 스냅샷"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "현황(스크린샷)"

    ws.merge_cells("A1:J1")
    ws["A1"] = f"BTS-600 일일 현황 — 화면 스크린샷 자동 판독  |  {when}  |  {len(rows)}개 회로"
    ws["A1"].font = Font(bold=True, size=12, color=NAVY)
    ws.row_dimensions[1].height = 22

    for j, c in enumerate(COLS, 1):
        cell = ws.cell(3, j, c)
        cell.fill = HEAD; cell.font = HEADF; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows):
        for j, key in enumerate(["circuit","battery","prog","status","cycle",
                                 "step","v","a","accu","dch"], 1):
            v = r.get(key, "")
            cell = ws.cell(4 + i, j, "무데이터" if nodata(v) else v)
            cell.border = BORDER; cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center")
            if nodata(v):
                cell.fill = NODATA
        if r.get("status") in ("CHA", "DCH"):
            ws.cell(4 + i, 4).fill = RUN

    widths = [9, 9, 11, 7, 7, 11, 9, 9, 11, 12]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
    ws.freeze_panes = "A4"

    Path(out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"저장: {out}  ({len(rows)}개 회로)")
    return out


if __name__ == "__main__":
    # 데모: 사용자가 보낸 BTS-600 화면 표를 비전으로 읽은 행들 (일부)
    rows = [
        {"battery":"Batt0007","circuit":"Circ0007","prog":"J2801new","status":"CHA","cycle":3,"step":"0:08:40","v":14.100,"a":0.482,"accu":286.186,"dch":597.404},
        {"battery":"Batt0008","circuit":"Circ0008","prog":"PJS","status":"CHA","cycle":0,"step":"8:26:52","v":15.999,"a":0.453,"accu":10.894,"dch":0.000},
        {"battery":"Batt0009","circuit":"Circ0009","prog":"PJS","status":"CHA","cycle":0,"step":"8:27:09","v":15.999,"a":0.502,"accu":11.525,"dch":0.000},
        {"battery":"Batt0013","circuit":"Circ0013","prog":"DOD175","status":"CHA","cycle":0,"step":"20:36:30","v":16.000,"a":0.540,"accu":17.682,"dch":-272.46},
        {"battery":"Batt0015","circuit":"Circ0015","prog":"GGH_C20","status":"CHA","cycle":0,"step":"17:58:56","v":15.999,"a":2.334,"accu":73.551,"dch":-272.57},
        {"battery":"Batt0023","circuit":"Circ0023","prog":"kjh_bo","status":"DCH","cycle":0,"step":"11:38:16","v":12.358,"a":-5.098,"accu":-11.185,"dch":59.332},
        {"battery":"Batt0027","circuit":"Circ0027","prog":"DoD50Joo","status":"CHA","cycle":21,"step":"0:18:06","v":16.577,"a":2.596,"accu":27.999,"dch":545.935},
        {"battery":"Batt0035","circuit":"Circ0035","prog":"HSHwater","status":"CHA","cycle":0,"step":"681:54:18","v":14.400,"a":0.108,"accu":92.258,"dch":0.000},
        {"battery":"Batt0036","circuit":"Circ0036","prog":"HSHwater","status":"CHA","cycle":0,"step":"681:53:45","v":14.399,"a":0.124,"accu":113.269,"dch":0.000},
        {"battery":"Batt0037","circuit":"Circ0037","prog":"PJS","status":"STO","cycle":0,"step":"0:00:00","v":0.001,"a":0.000,"accu":29.482,"dch":43.927},
    ]
    build(rows, when="2026-07-22 08:27 (데모: 화면 일부)")
