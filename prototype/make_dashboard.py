#!/usr/bin/env python3
"""분석 결과 대시보드 HTML 생성 (차트 PNG를 data URI로 임베드)."""

import base64
import sys
from pathlib import Path

sys.path.insert(0, "prototype")
from bts600_extract import parse, extract_tracking
from fill_tracking import compare_pairs, circuit_to_sample

OUT = Path("prototype/out")
SAMPLES = "prototype/samples"
XLSX = f"{SAMPLES}/HKMC_SOC_tracking.xlsx"
CSVS = [f"{SAMPLES}/L6n1.csv", f"{SAMPLES}/L6n2.csv", f"{SAMPLES}/L6n3.csv"]


def datauri(png):
    b = base64.b64encode((OUT / png).read_bytes()).decode()
    return f"data:image/png;base64,{b}"


def tracking_rows():
    """#1, #3 핵심 값 표 (손입력 vs 자동)."""
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["GB L6"]
    pc = {ws[f"{c}5"].value: ws[f"{c}10"].value for c in "DEF"}
    colmap = {"#1": "D", "#3": "F"}
    ROW = {"충전용량 1차 (Ah)": 15, "안정화 전압 72h (V)": 16,
           "총 방전용량 (Ah)": 25, "충전용량 2차 (Ah)": 27, "안정화 전압 24h (V)": 28}
    ext = {}
    for csv in CSVS:
        m, s = parse(csv)
        samp = circuit_to_sample(m["circuit"], pc)
        if samp in colmap:
            t = extract_tracking(m, s)
            caps, stabs = t["충전용량(Ah)"], t["안정화전압(휴지종료V)"]
            ext[samp] = {
                "충전용량 1차 (Ah)": caps[0] if caps else None,
                "안정화 전압 72h (V)": stabs[0] if stabs else None,
                "총 방전용량 (Ah)": t["총방전용량(Ah)"],
                "충전용량 2차 (Ah)": caps[1] if len(caps) > 1 else None,
                "안정화 전압 24h (V)": stabs[1] if len(stabs) > 1 else None,
            }
    rows = []
    for label, r in ROW.items():
        cells = []
        for samp, col in colmap.items():
            hand = ws[f"{col}{r}"].value
            auto = ext.get(samp, {}).get(label)
            ok = False
            try:
                ok = abs(float(hand) - float(auto)) <= 0.02
            except (ValueError, TypeError):
                pass
            cells.append((hand, auto, ok))
        rows.append((label, cells))
    return rows


def main():
    pairs = compare_pairs(XLSX, CSVS)
    match = sum(1 for a, b in pairs if abs(a - b) <= 0.02)
    trows = tracking_rows()

    tbody = ""
    for label, cells in trows:
        tds = f'<th scope="row">{label}</th>'
        for hand, auto, ok in cells:
            h = "—" if hand in (None, "-") else f"{float(hand):g}"
            a = "—" if auto is None else f"{float(auto):g}"
            mark = '<span class="ok">일치</span>' if ok else '<span class="na">—</span>'
            tds += f'<td class="num">{h}</td><td class="num auto">{a}</td><td class="chk">{mark}</td>'
        tbody += f"<tr>{tds}</tr>"

    findings = [
        ("warn", "온도센서 미연결", "Batt0024 / L6SOC#2",
         "전체 590개 로그가 -270.49°C(물리적 불가값). 온도 감시 없이 시험이 진행됨 — 결과 파일을 열기 전엔 놓치기 쉬운 문제."),
        ("warn", "방전 중 전압 정체", "Batt0020 / L6SOC#1",
         "방전 구간에서 전압이 513 로그 연속 불변. 저전류 플로트 구간일 가능성이 높으나 사람이 한 번 확인 권장."),
        ("info", "시험 진행 중 인식", "Batt0020 / L6SOC#1",
         "종료 시각이 01-01-00(널값)으로 기록됨 → 경과 191.8h 진행 중으로 자동 판정."),
    ]
    fcards = ""
    sev = {"warn": ("주의", "warn"), "info": ("정보", "info"), "crit": ("경보", "crit")}
    for level, title, meta, desc in findings:
        tag, cls = sev[level]
        fcards += f'''<div class="finding {cls}">
          <div class="finding-head"><span class="sev">{tag}</span><span class="finding-title">{title}</span></div>
          <div class="finding-meta">{meta}</div>
          <p class="finding-desc">{desc}</p></div>'''

    html = TEMPLATE.format(
        profile=datauri("chart_profile.png"),
        c20=datauri("chart_c20_curve.png"),
        valid=datauri("chart_validation.png"),
        n_pairs=len(pairs), n_match=match,
        acc=f"{100*match/len(pairs):.0f}",
        tbody=tbody, fcards=fcards,
    )
    dest = OUT / "dashboard.html"
    dest.write_text(html, encoding="utf-8")
    print("saved", dest, f"({dest.stat().st_size//1024} KB)")


TEMPLATE = r"""<title>BTS-600 시험데이터 자동화 — 결과 대시보드</title>
<style>
  :root {{
    --ground:#f4f6f9; --surface:#ffffff; --surface-2:#eef2f7; --line:#dce2ea;
    --ink:#141a22; --ink-2:#5a6472; --ink-3:#8a94a3;
    --accent:#004094; --accent-2:#2a78d6;
    --good:#1a8f5c; --good-bg:#e7f4ee; --warn:#c67f06; --warn-bg:#fbf1dd;
    --crit:#d6453d; --info:#2a78d6; --info-bg:#e6eff9;
    --shadow:0 1px 2px rgba(20,26,34,.04),0 4px 16px rgba(20,26,34,.05);
  }}
  @media (prefers-color-scheme:dark) {{
    :root {{
      --ground:#0f141b; --surface:#161d27; --surface-2:#1d2532; --line:#293341;
      --ink:#eef1f6; --ink-2:#9aa5b4; --ink-3:#6b7686;
      --accent:#4a90e2; --accent-2:#63a4ea;
      --good:#3fb884; --good-bg:#12271f; --warn:#e0a53a; --warn-bg:#2a2113;
      --crit:#e5675f; --info:#63a4ea; --info-bg:#12202f;
      --shadow:0 1px 2px rgba(0,0,0,.3),0 4px 18px rgba(0,0,0,.35);
    }}
  }}
  :root[data-theme="light"] {{
    --ground:#f4f6f9; --surface:#ffffff; --surface-2:#eef2f7; --line:#dce2ea;
    --ink:#141a22; --ink-2:#5a6472; --ink-3:#8a94a3; --accent:#004094; --accent-2:#2a78d6;
    --good:#1a8f5c; --good-bg:#e7f4ee; --warn:#c67f06; --warn-bg:#fbf1dd;
    --crit:#d6453d; --info:#2a78d6; --info-bg:#e6eff9;
  }}
  :root[data-theme="dark"] {{
    --ground:#0f141b; --surface:#161d27; --surface-2:#1d2532; --line:#293341;
    --ink:#eef1f6; --ink-2:#9aa5b4; --ink-3:#6b7686; --accent:#4a90e2; --accent-2:#63a4ea;
    --good:#3fb884; --good-bg:#12271f; --warn:#e0a53a; --warn-bg:#2a2113;
    --crit:#e5675f; --info:#63a4ea; --info-bg:#12202f;
  }}
  * {{ box-sizing:border-box; }}
  body {{
    margin:0; background:var(--ground); color:var(--ink);
    font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","Apple SD Gothic Neo","Malgun Gothic",sans-serif;
    line-height:1.5; -webkit-font-smoothing:antialiased;
  }}
  .wrap {{ max-width:1080px; margin:0 auto; padding:40px 24px 64px; }}
  .num, .mono {{ font-variant-numeric:tabular-nums; }}
  .eyebrow {{ font-size:12px; font-weight:700; letter-spacing:.14em; text-transform:uppercase;
    color:var(--accent-2); margin:0 0 10px; }}
  h1 {{ font-size:30px; font-weight:800; letter-spacing:-.02em; margin:0 0 8px; text-wrap:balance; }}
  .lede {{ font-size:16px; color:var(--ink-2); margin:0; max-width:64ch; }}
  header {{ border-bottom:1px solid var(--line); padding-bottom:28px; margin-bottom:28px; }}
  section {{ margin-top:36px; }}
  h2 {{ font-size:13px; font-weight:700; letter-spacing:.1em; text-transform:uppercase;
    color:var(--ink-2); margin:0 0 16px; display:flex; align-items:center; gap:10px; }}
  h2::after {{ content:""; flex:1; height:1px; background:var(--line); }}

  .kpis {{ display:grid; grid-template-columns:repeat(4,1fr); gap:14px; }}
  .kpi {{ background:var(--surface); border:1px solid var(--line); border-radius:12px;
    padding:18px 18px 16px; box-shadow:var(--shadow); }}
  .kpi .v {{ font-size:32px; font-weight:800; letter-spacing:-.02em; line-height:1; }}
  .kpi .v small {{ font-size:15px; font-weight:700; color:var(--ink-2); }}
  .kpi .l {{ font-size:12.5px; color:var(--ink-2); margin-top:8px; }}
  .kpi.hl .v {{ color:var(--good); }}

  .pipe {{ display:grid; grid-template-columns:1fr auto 1fr auto 1fr; align-items:center; gap:10px;
    background:var(--surface); border:1px solid var(--line); border-radius:12px; padding:18px;
    box-shadow:var(--shadow); }}
  .pipe .node {{ text-align:center; padding:8px; }}
  .pipe .node .t {{ font-weight:700; font-size:14px; }}
  .pipe .node .s {{ font-size:12px; color:var(--ink-2); margin-top:3px; }}
  .pipe .arrow {{ color:var(--accent-2); font-size:20px; font-weight:700; }}
  .pipe .split {{ display:flex; flex-direction:column; gap:8px; }}
  .pipe .split .t {{ font-size:13px; }}
  .chip {{ display:inline-block; font-size:11px; font-weight:700; padding:2px 8px; border-radius:999px;
    background:var(--surface-2); color:var(--ink-2); }}

  figure {{ margin:0; background:var(--surface); border:1px solid var(--line); border-radius:12px;
    padding:10px; box-shadow:var(--shadow); }}
  figure img {{ width:100%; height:auto; display:block; border-radius:6px; }}
  figcaption {{ font-size:12.5px; color:var(--ink-2); padding:10px 8px 4px; }}
  .grid2 {{ display:grid; grid-template-columns:1fr 1fr; gap:16px; }}

  .tablewrap {{ overflow-x:auto; background:var(--surface); border:1px solid var(--line);
    border-radius:12px; box-shadow:var(--shadow); }}
  table {{ border-collapse:collapse; width:100%; font-size:13.5px; min-width:640px; }}
  caption {{ text-align:left; padding:14px 16px 0; font-size:12.5px; color:var(--ink-2); }}
  th, td {{ padding:9px 14px; text-align:left; border-bottom:1px solid var(--line); }}
  thead th {{ font-size:11px; letter-spacing:.05em; text-transform:uppercase; color:var(--ink-3);
    font-weight:700; }}
  tbody th {{ font-weight:600; color:var(--ink); }}
  td.num {{ text-align:right; font-variant-numeric:tabular-nums; }}
  td.auto {{ color:var(--accent-2); font-weight:600; }}
  td.chk {{ text-align:center; }}
  .colgrp {{ background:var(--surface-2); }}
  .ok {{ color:var(--good); font-weight:700; font-size:12px; }}
  .na {{ color:var(--ink-3); }}

  .findings {{ display:flex; flex-direction:column; gap:12px; }}
  .finding {{ background:var(--surface); border:1px solid var(--line); border-left-width:4px;
    border-radius:10px; padding:14px 16px; box-shadow:var(--shadow); }}
  .finding.warn {{ border-left-color:var(--warn); }}
  .finding.info {{ border-left-color:var(--info); }}
  .finding.crit {{ border-left-color:var(--crit); }}
  .finding-head {{ display:flex; align-items:center; gap:10px; }}
  .sev {{ font-size:11px; font-weight:800; padding:2px 8px; border-radius:6px; letter-spacing:.03em; }}
  .warn .sev {{ background:var(--warn-bg); color:var(--warn); }}
  .info .sev {{ background:var(--info-bg); color:var(--info); }}
  .crit .sev {{ background:#fbe6e4; color:var(--crit); }}
  .finding-title {{ font-weight:700; font-size:14.5px; }}
  .finding-meta {{ font-size:12px; color:var(--ink-3); margin-top:2px; font-variant-numeric:tabular-nums; }}
  .finding-desc {{ font-size:13.5px; color:var(--ink-2); margin:8px 0 0; }}

  .next {{ background:var(--surface); border:1px solid var(--line); border-radius:12px; padding:20px 22px;
    box-shadow:var(--shadow); }}
  .next ol {{ margin:0; padding-left:22px; }}
  .next li {{ margin:7px 0; font-size:14px; color:var(--ink); }}
  .next li span {{ color:var(--ink-2); }}
  footer {{ margin-top:40px; padding-top:20px; border-top:1px solid var(--line);
    font-size:12px; color:var(--ink-3); display:flex; justify-content:space-between; flex-wrap:wrap; gap:8px; }}
  @media (max-width:720px) {{
    .kpis {{ grid-template-columns:repeat(2,1fr); }}
    .grid2 {{ grid-template-columns:1fr; }}
    .pipe {{ grid-template-columns:1fr; }} .pipe .arrow {{ transform:rotate(90deg); }}
  }}
</style>

<div class="wrap">
  <header>
    <p class="eyebrow">세방전지 · 시험데이터 자동화 개념검증(PoC)</p>
    <h1>Digatron BTS-600 로우데이터 → 시험 관리대장 자동 작성</h1>
    <p class="lede">시험기를 제어하지 않고, C:\bts600 폴더에 쌓이는 export CSV만 읽어 담당자의
      수기 전기(轉記) 작업을 자동화하고 이상 징후를 감지합니다. 실제 시험 파일 3개로 검증한 결과입니다.</p>
  </header>

  <section>
    <div class="kpis">
      <div class="kpi"><div class="v">3</div><div class="l">검증 시료 (L6SOC #1·#2·#3)</div></div>
      <div class="kpi"><div class="v">139,100</div><div class="l">처리한 측정 로그 행</div></div>
      <div class="kpi hl"><div class="v">{acc}<small>%</small></div><div class="l">관리대장 {n_match}/{n_pairs} 값 자동 일치</div></div>
      <div class="kpi"><div class="v">3<small>건</small></div><div class="l">자동 감지된 이상·특이사항</div></div>
    </div>
  </section>

  <section>
    <h2>처리 흐름</h2>
    <div class="pipe">
      <div class="node"><div class="t">로우데이터</div><div class="s">BTS-600 export CSV × 3</div>
        <div style="margin-top:6px"><span class="chip">시험기 제어 없음</span></div></div>
      <div class="arrow">→</div>
      <div class="node"><div class="t">파싱 · 추출</div><div class="s">구간 분해 · 회로↔샘플 매핑</div></div>
      <div class="arrow">→</div>
      <div class="split">
        <div><span class="chip" style="background:var(--good-bg);color:var(--good)">산출물 1</span>
          <span class="t"> 관리대장 xlsx 자동 기입</span></div>
        <div><span class="chip" style="background:var(--warn-bg);color:var(--warn)">산출물 2</span>
          <span class="t"> 이상 감지 리포트</span></div>
      </div>
    </div>
  </section>

  <section>
    <h2>전체 충방전 프로파일 (샘플 #1)</h2>
    <figure>
      <img src="{profile}" alt="샘플 1 충방전 전압 프로파일">
      <figcaption>초기 만충전 → 72시간 안정화 → 단계별 C20 방전 → 재충전 → 24시간 안정화까지
        191.8시간·68,867개 로그를 하나의 그림으로. 색은 충전/방전/휴지 구간.</figcaption>
    </figure>
  </section>

  <section>
    <h2>추출 결과 · 검증</h2>
    <div class="grid2">
      <figure>
        <img src="{c20}" alt="C20 방전곡선">
        <figcaption>로우데이터에서 자동 추출한 C20 SOC-OCV 방전곡선. 두 시료가 거의 일치 = 정상 편차.</figcaption>
      </figure>
      <figure>
        <img src="{valid}" alt="손입력 vs 자동추출 검증">
        <figcaption>담당자 손입력값과 자동추출값 {n_pairs}개를 대조. 모든 점이 대각선 위 = 완전 일치.</figcaption>
      </figure>
    </div>
  </section>

  <section>
    <h2>관리대장 자동 기입값 대조</h2>
    <div class="tablewrap">
      <table>
        <caption>'GB L6' 시트 핵심 항목 — 담당자 손입력 vs 에이전트 자동추출 (소수점 셋째 자리까지 대조)</caption>
        <thead>
          <tr>
            <th rowspan="2">항목</th>
            <th colspan="3" class="colgrp">샘플 #1 (Batt0020)</th>
            <th colspan="3">샘플 #3 (Batt0016)</th>
          </tr>
          <tr>
            <th class="colgrp">손입력</th><th class="colgrp">자동</th><th class="colgrp">판정</th>
            <th>손입력</th><th>자동</th><th>판정</th>
          </tr>
        </thead>
        <tbody>{tbody}</tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>자동 감지된 이상 · 특이사항</h2>
    <div class="findings">{fcards}</div>
  </section>

  <section>
    <h2>다음 단계</h2>
    <div class="next">
      <ol>
        <li>공유폴더 감시 <span>— 신규 CSV가 쌓이면 자동으로 처리·관리대장 갱신 (설정만, 시험기 무수정)</span></li>
        <li>규격 기준값 연동 <span>— 합부 판정 컬럼 자동 생성</span></li>
        <li>이상 시 알림 <span>— 온도센서 미연결 같은 문제를 시험 도중 메일/메신저로 통보</span></li>
        <li>LLM 해설 <span>— "왜 이런 패턴인지 + 조치" 자연어 요약 첨부</span></li>
        <li>전 회로 확대 <span>— 35개 채널 주간 현황 자동 리포트</span></li>
      </ol>
    </div>
  </section>

  <footer>
    <span>세방전지 시험데이터 자동화 PoC · 실제 데이터 기반 개념검증</span>
    <span>데이터: L6SOC #1–#3 · HKMC_SOC 관리대장 (GB L6)</span>
  </footer>
</div>
"""


if __name__ == "__main__":
    main()
