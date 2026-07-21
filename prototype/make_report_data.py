#!/usr/bin/env python3
"""기안서 생성용 데이터 번들 생성 (샘플정보 + 시험결과 + 판정 → JSON)."""

import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from bts600_extract import parse, extract_tracking
from bts600_analyzer import parse as aparse, analyze
from judge import judge
from fill_tracking import circuit_to_sample

XLSX = "prototype/samples/HKMC_SOC_tracking.xlsx"
CSVS = ["prototype/samples/L6n1.csv", "prototype/samples/L6n2.csv", "prototype/samples/L6n3.csv"]
MODEL = "GB L6"
SOC = [100, 90, 80, 70, 60, 40, 20, 0]


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["GB L6"]
    pc = {ws[f"{c}5"].value: ws[f"{c}10"].value for c in "DEF"}
    info_rows = {"weight": 6, "voltage": 7, "resistance": 8, "mcca": 9,
                 "pc": 10, "start": 11}
    col = {"#1": "D", "#2": "E", "#3": "F"}

    samples = []
    for csv in CSVS:
        m, steps = parse(csv)
        samp = circuit_to_sample(m["circuit"], pc)
        if not samp:
            continue
        t = extract_tracking(m, steps)
        am, arows = aparse(csv)
        findings, summ = analyze(am, arows)
        verdict, judgments = judge(MODEL, t, findings)
        c = col[samp]
        start = ws[f"{c}11"].value
        samples.append({
            "sample": samp,
            "battery": m["battery"], "circuit": m["circuit"], "section": m["section"],
            "program": m["program"],
            "weight": ws[f"{c}6"].value, "voltage_init": ws[f"{c}7"].value,
            "resistance": ws[f"{c}8"].value, "mcca": ws[f"{c}9"].value,
            "pc": ws[f"{c}10"].value,
            "start": start.strftime("%Y-%m-%d") if hasattr(start, "strftime") else str(start),
            "running": summ.get("running"), "duration_h": summ.get("duration_h"),
            "charge_caps": t["충전용량(Ah)"],
            "total_discharge": t["총방전용량(Ah)"],
            "stab": t["안정화전압(휴지종료V)"],
            "soc_curve": t["SOC구간 OCV(V)"],
            "verdict": verdict,
            "judgments": [{"item": j.item, "value": j.value, "limit": j.limit, "ok": j.ok}
                          for j in judgments],
            "notable": [f"[{f.level}] {f.message}" for f in findings if f.level in ("ALARM", "WARN")],
        })

    bundle = {
        "model": MODEL, "spec": ws["D4"].value, "soc_labels": SOC,
        "samples": samples,
        "n": len(samples),
        "n_pass": sum(1 for s in samples if s["verdict"] == "PASS"),
    }
    out = Path("prototype/out/report_data.json")
    out.write_text(json.dumps(bundle, ensure_ascii=False, indent=2), encoding="utf-8")
    print("saved", out, f"({len(samples)} samples)")


if __name__ == "__main__":
    main()
