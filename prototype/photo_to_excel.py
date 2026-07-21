#!/usr/bin/env python3
"""
샘플 라벨 사진 → 엑셀 자동 기입 (계측기 연동 없이 비전 OCR)

흐름:
  라벨 사진  →  비전 모델(Claude)이 손글씨 판독 → 구조화 JSON  →  엑셀 기입
             (신뢰도 낮은 필드는 노란색으로 표시 → 담당자 확인)

production: extract_with_claude(image_path) 가 Claude API(vision)를 호출해 JSON을 받는다.
demo:       vision 판독 결과를 JSON으로 받아 그대로 엑셀에 기입한다.
"""

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY = "004094"
HEAD = PatternFill("solid", fgColor=NAVY)
HEADF = Font(color="FFFFFF", bold=True, size=10)
CHECK = PatternFill("solid", fgColor="FFF2CC")   # 확인 필요(신뢰도 낮음)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 엑셀 컬럼 = 라벨에서 뽑을 항목
FIELDS = ["샘플번호", "중량(g)", "전압(V)", "내부저항(mΩ)", "MCCA", "프로그램", "시험명/요청", "일자"]


VISION_PROMPT = """이 사진은 배터리에 붙은 손글씨 라벨입니다.
다음 항목을 읽어 JSON으로만 답하세요. 각 항목에 value 와 confidence(high/med/low)를 넣으세요.
읽을 수 없으면 value=null, confidence="low".
항목: 샘플번호, 중량(g), 전압(V), 내부저항(mΩ), MCCA, 프로그램, 시험명/요청, 일자
형식: {"샘플번호":{"value":"#5","confidence":"high"}, ...}"""


def extract_with_claude(image_path):
    """
    production 구현 — Claude API(vision)로 라벨 판독.
    실제 사용 시 아래 주석 코드를 활성화 (anthropic SDK + API 키 필요).

        import anthropic, base64
        client = anthropic.Anthropic()
        img = base64.standard_b64encode(Path(image_path).read_bytes()).decode()
        msg = client.messages.create(
            model="claude-opus-4-8", max_tokens=1024,
            messages=[{"role":"user","content":[
                {"type":"image","source":{"type":"base64","media_type":"image/png","data":img}},
                {"type":"text","text":VISION_PROMPT}]}])
        return json.loads(msg.content[0].text)
    """
    raise NotImplementedError("데모에서는 vision 판독 결과 JSON을 직접 전달")


def fill_excel(records, out="prototype/out/sample_info.xlsx"):
    """records: [{필드:{value,confidence}}, ...] → 엑셀 기입."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "샘플 정보"

    ws.merge_cells("A1:H1")
    ws["A1"] = "샘플 정보 (라벨 사진 자동 판독)"
    ws["A1"].font = Font(bold=True, size=13, color=NAVY)
    ws.row_dimensions[1].height = 22

    for j, f in enumerate(FIELDS, 1):
        c = ws.cell(3, j, f)
        c.fill = HEAD; c.font = HEADF; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

    n_check = 0
    for i, rec in enumerate(records):
        for j, f in enumerate(FIELDS, 1):
            cell = ws.cell(4 + i, j)
            item = rec.get(f, {})
            v = item.get("value")
            cell.value = "" if v is None else v
            cell.border = BORDER
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center")
            if item.get("confidence") == "low" or v is None:
                cell.fill = CHECK            # 확인 필요 강조
                n_check += 1

    ws.cell(4 + len(records) + 1, 1,
            f"노란색 = 판독 신뢰도 낮음, 담당자 확인 필요 ({n_check}칸)").font = \
        Font(italic=True, size=10, color="C67F06")

    widths = [10, 10, 9, 12, 8, 10, 16, 12]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

    Path(out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"저장: {out}  ({len(records)}개 샘플, 확인필요 {n_check}칸)")
    return out


if __name__ == "__main__":
    # 데모: 업로드된 라벨 사진(#5)을 비전으로 판독한 결과
    demo = [{
        "샘플번호":   {"value": "#5",   "confidence": "high"},
        "중량(g)":    {"value": None,   "confidence": "low"},   # 흐릿 → 확인필요
        "전압(V)":    {"value": 12.88,  "confidence": "high"},
        "내부저항(mΩ)": {"value": 4.76, "confidence": "high"},
        "MCCA":       {"value": None,   "confidence": "low"},   # 라벨에 없음
        "프로그램":    {"value": "PJS",  "confidence": "high"},
        "시험명/요청": {"value": None,   "confidence": "low"},   # 손글씨 판독 애매
        "일자":       {"value": None,   "confidence": "low"},
    }]
    fill_excel(demo)
