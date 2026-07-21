#!/usr/bin/env python3
"""
일일 시험현황 엑셀 자동 기입 (Type A)

폴더의 BTS-600 CSV(40개 회로)를 읽어, 각 회로의 '중간결과 스냅샷'을
엑셀 한 장에 회로당 1행으로 자동 기입한다.
  기입 항목: 상태 · 경과시간 · 현재전압 · 전류 · 충전용량 · 방전용량 · 사이클 · 판정

    python3 daily_report.py <CSV폴더> [모델] [날짜라벨]
"""

import csv
import io
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).parent))
from bts600_analyzer import parse as aparse, analyze
from bts600_extract import parse as eparse, extract_tracking
from judge import judge

NAVY = "004094"
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
RUN_FILL = PatternFill("solid", fgColor="FBF1DD")     # 진행중
PASS_FILL = PatternFill("solid", fgColor="E7F4EE")
FAIL_FILL = PatternFill("solid", fgColor="FBE6E4")
WARN_FILL = PatternFill("solid", fgColor="FFF6E5")
THIN = Side(style="thin", color="DCE2EA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = ["회로", "시료", "시험구간", "프로그램", "상태", "경과(h)",
        "현재전압(V)", "전류(A)", "충전용량(Ah)", "방전용량(Ah)",
        "사이클", "이상", "판정", "갱신"]


def snapshot(csv_path):
    """마지막 유효 데이터 행 = 현재 중간결과 스냅샷 + 누적 최대값."""
    txt = open(csv_path, encoding="utf-8", errors="replace").read()
    r = list(csv.reader(io.StringIO(txt)))
    hi = next(i for i, x in enumerate(r) if x and x[0].strip() == "Step" and "Status" in x)
    col = {n.strip(): j for j, n in enumerate(r[hi])}

    def g(row, name):
        j = col.get(name)
        return row[j] if j is not None and j < len(row) else ""

    def f(v):
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    last, max_cha, max_dch, max_cycle = None, 0.0, 0.0, 0
    for row in r[hi + 2:]:
        if not row or not row[0].strip() or g(row, "Status").strip() == "":
            continue
        last = row
        c, d = f(g(row, "AhCha")), f(g(row, "AhDch"))
        if c:
            max_cha = max(max_cha, c)
        if d:
            max_dch = max(max_dch, d)
        cy = f(g(row, "Cycle"))
        if cy:
            max_cycle = max(max_cycle, int(cy))
    st = {"CHA": "충전", "DCH": "방전", "PAU": "휴지"}.get(g(last, "Status").strip(), "-")
    pt = g(last, "Program time").split(":")
    ph = int(pt[0]) + int(pt[1]) / 60 if len(pt) >= 2 else 0
    return {
        "status": st, "prog_h": round(ph, 1),
        "voltage": f(g(last, "Voltage")), "current": f(g(last, "Current")),
        "cha": round(max_cha, 3), "dch": round(max_dch, 3), "cycle": max_cycle,
    }


def build(folder, model, datelabel):
    csvs = sorted(Path(folder).glob("*.csv"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "일일 시험현황"

    # 제목
    ws.merge_cells("A1:N1")
    t = ws["A1"]
    t.value = f"일일 시험현황  |  모델 {model}  |  {datelabel}  |  {len(csvs)}개 회로"
    t.font = Font(bold=True, size=13, color=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # 헤더
    for j, name in enumerate(COLS, 1):
        c = ws.cell(3, j, name)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[3].height = 20

    row = 4
    n_run = n_warn = 0
    for path in csvs:
        s = snapshot(path)
        m, steps = eparse(path)
        tracking = extract_tracking(m, steps)
        am, arows = aparse(path)
        findings, summ = analyze(am, arows)
        verdict, _ = judge(model, tracking, findings)
        notable = [f for f in findings if f.level in ("ALARM", "WARN")]
        running = summ.get("running")
        if running:
            n_run += 1
        if notable:
            n_warn += 1

        vals = [
            m["circuit"], m["battery"], m["section"], m["program"], s["status"],
            s["prog_h"], s["voltage"], s["current"], s["cha"], s["dch"],
            s["cycle"], ("⚠ " + str(len(notable))) if notable else "",
            {"PASS": "합격", "FAIL": "불합격", "HOLD": "보류", "N/A": "-"}[verdict],
            datelabel,
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row, j, v)
            c.border = BORDER
            c.font = Font(size=10)
            if j >= 6 and j <= 11:
                c.alignment = Alignment(horizontal="right")
                c.number_format = "0.0" if j in (6, 7, 8) else "0.000" if j in (9, 10) else "0"
            else:
                c.alignment = Alignment(horizontal="center")
        # 색상: 진행중 상태칸, 판정칸
        if running:
            ws.cell(row, 5).fill = RUN_FILL
        if verdict == "PASS":
            ws.cell(row, 13).fill = PASS_FILL
        elif verdict == "FAIL":
            ws.cell(row, 13).fill = FAIL_FILL
        if notable:
            ws.cell(row, 12).fill = WARN_FILL
        row += 1

    # 요약 행
    ws.cell(row + 1, 1, f"진행중 {n_run} · 이상감지 {n_warn} · 총 {len(csvs)}개 회로")
    ws.cell(row + 1, 1).font = Font(italic=True, size=10, color="5A6472")

    widths = [9, 9, 11, 12, 7, 8, 11, 9, 12, 12, 7, 7, 8, 11]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
    ws.freeze_panes = "A4"

    out = Path("prototype/out/daily_status.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"저장: {out}  ({len(csvs)}개 회로, 진행중 {n_run}, 이상 {n_warn})")
    return out


if __name__ == "__main__":
    folder = sys.argv[1] if len(sys.argv) > 1 else "prototype/samples"
    model = sys.argv[2] if len(sys.argv) > 2 else "GB L6"
    datelabel = sys.argv[3] if len(sys.argv) > 3 else "2025-11-28 08:00"
    build(folder, model, datelabel)
