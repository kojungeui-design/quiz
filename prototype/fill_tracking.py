#!/usr/bin/env python3
"""
BTS-600 로우데이터 → 실제 관리대장 xlsx 자동 채움 + 검증

- 로우데이터 CSV 3개를 읽어 관리대장 값 추출
- 회로번호 → 샘플 열(D/E/F) 자동 매핑
- 관리대장 사본에 값 기입(자동 채운 셀은 초록 하이라이트)
- 원본(담당자 손입력)과 대조해 일치/불일치 검증 리포트
"""

import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font

from bts600_extract import parse, extract_tracking

GREEN = PatternFill("solid", fgColor="C6EFCE")   # 자동 채움
RED = PatternFill("solid", fgColor="FFC7CE")     # 불일치

# 관리대장 'GB L6' 시트 값 행(row) → 추출 항목 매핑
# 열: D=#1, E=#2, F=#3
SAMPLE_COL = {"#1": "D", "#2": "E", "#3": "F"}
SOC_ROWS = [17, 18, 19, 20, 21, 22, 23, 24]  # SOC 1.0,0.9,0.8,0.7,0.6,0.4,0.2,0
ROW = {"충전용량1": 15, "안정화72h": 16, "총방전용량": 25, "충전용량2": 27, "안정화24h": 28}


def circuit_to_sample(circuit, pc_map):
    """Circ0020 → PC번호 매칭 → 샘플 열."""
    num = circuit.replace("Circ", "").lstrip("0")
    for sample, pc in pc_map.items():
        # PC "4-20" 의 뒷자리가 회로번호 끝자리와 일치
        if pc.split("-")[-1] == num:
            return sample
    return None


def main(xlsx_in, csv_paths, sheet="GB L6"):
    out_path = Path("prototype/out/HKMC_SOC_autofilled.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(xlsx_in, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb[sheet]

    # PC번호 맵 읽기 (10행)
    pc_map = {ws[f"{c}5"].value: ws[f"{c}10"].value for c in "DEF"}
    print("PC번호 맵:", pc_map)

    # 원본 손입력값 백업 (검증용)
    orig = {}
    for name, r in ROW.items():
        for c in "DEF":
            orig[(name, c)] = ws[f"{c}{r}"].value
    for i, r in enumerate(SOC_ROWS):
        for c in "DEF":
            orig[(f"soc{i}", c)] = ws[f"{c}{r}"].value

    match, mismatch, filled = 0, 0, 0
    report = []

    for csv in csv_paths:
        meta, steps = parse(csv)
        t = extract_tracking(meta, steps)
        sample = circuit_to_sample(meta["circuit"], pc_map)
        if not sample:
            print(f"  [스킵] {meta['circuit']} → 매칭되는 샘플 없음")
            continue
        col = SAMPLE_COL[sample]
        print(f"\n{Path(csv).name}: {meta['circuit']} → 샘플 {sample} (열 {col})")

        vals = {}
        caps = t["충전용량(Ah)"]
        stabs = t["안정화전압(휴지종료V)"]
        if caps:
            vals[ROW["충전용량1"]] = caps[0]
        if len(caps) > 1:
            vals[ROW["충전용량2"]] = caps[1]
        if stabs:
            vals[ROW["안정화72h"]] = stabs[0]
        if len(stabs) > 1:
            vals[ROW["안정화24h"]] = stabs[1]
        vals[ROW["총방전용량"]] = t["총방전용량(Ah)"]
        for i, v in enumerate(t["SOC구간 OCV(V)"]):
            if i < len(SOC_ROWS):
                vals[SOC_ROWS[i]] = v

        for r, v in vals.items():
            cell = ws[f"{col}{r}"]
            hand = orig.get(_row_key(r, col))
            # 비교 (손입력이 '-' 또는 없으면 채우기만)
            ok = _close(hand, v)
            cell.value = v
            cell.fill = GREEN if (hand in (None, "-") or ok) else RED
            filled += 1
            if hand not in (None, "-"):
                if ok:
                    match += 1
                else:
                    mismatch += 1
                    report.append(f"  ✗ {col}{r}: 손입력 {hand} ≠ 자동 {v}")

    wb.save(out_path)
    print("\n" + "=" * 60)
    print(f"자동 기입: {filled}개 셀")
    print(f"원본 대조: 일치 {match} / 불일치 {mismatch}")
    for line in report:
        print(line)
    acc = 100 * match / (match + mismatch) if (match + mismatch) else 0
    print(f"정확도: {acc:.1f}%")
    print(f"저장: {out_path}")
    return match, mismatch


def compare_pairs(xlsx_in, csv_paths, sheet="GB L6"):
    """(손입력, 자동추출) 쌍 리스트 반환 — 검증 차트용. 파일 수정 없음."""
    wb = openpyxl.load_workbook(xlsx_in, data_only=True)
    ws = wb[sheet]
    pc_map = {ws[f"{c}5"].value: ws[f"{c}10"].value for c in "DEF"}
    pairs = []
    for csv in csv_paths:
        meta, steps = parse(csv)
        t = extract_tracking(meta, steps)
        sample = circuit_to_sample(meta["circuit"], pc_map)
        if not sample:
            continue
        col = SAMPLE_COL[sample]
        vals = {}
        caps, stabs = t["충전용량(Ah)"], t["안정화전압(휴지종료V)"]
        if caps:
            vals[ROW["충전용량1"]] = caps[0]
        if len(caps) > 1:
            vals[ROW["충전용량2"]] = caps[1]
        if stabs:
            vals[ROW["안정화72h"]] = stabs[0]
        if len(stabs) > 1:
            vals[ROW["안정화24h"]] = stabs[1]
        vals[ROW["총방전용량"]] = t["총방전용량(Ah)"]
        for i, v in enumerate(t["SOC구간 OCV(V)"]):
            if i < len(SOC_ROWS):
                vals[SOC_ROWS[i]] = v
        for r, v in vals.items():
            hand = ws[f"{col}{r}"].value
            try:
                pairs.append((float(hand), float(v)))
            except (ValueError, TypeError):
                pass
    return pairs


def _row_key(r, col):
    for name, rr in ROW.items():
        if rr == r:
            return (name, col)
    if r in SOC_ROWS:
        return (f"soc{SOC_ROWS.index(r)}", col)
    return (str(r), col)


def _close(a, b, tol=0.02):
    try:
        return abs(float(a) - float(b)) <= tol
    except (ValueError, TypeError):
        return False


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "prototype/samples/HKMC_SOC_tracking.xlsx"
    csvs = sys.argv[2:] or ["prototype/samples/L6n1.csv",
                            "prototype/samples/L6n2.csv",
                            "prototype/samples/L6n3.csv"]
    main(xlsx, csvs)
