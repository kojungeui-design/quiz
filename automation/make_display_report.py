#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================================
#  make_display_report.py  —  BTS-600 CSV  →  "측정값 표시" 엑셀 리포트
# ---------------------------------------------------------------------
#  BTS-600 이 export 한 CSV(Measuring data list) 를 그대로 엑셀로 옮기되,
#  시험의 '핵심 측정 지점' 행을 노란색으로 표시(highlight) 한다.
#  → 사용자가 손으로 만들던 "○○ - 표시.xlsx" 와 동일한 형식.
#
#  값을 요약/가공하지 않고 원본 그대로 옮기므로 값이 정확하다.
#  (기존 요약 리포트가 "너무 간단하고 값이 맞는지 모르겠다" 던 문제 해결)
#
#  사용법:
#     python make_display_report.py  파일.csv                → 파일 - 표시.xlsx
#     python make_display_report.py  폴더                    → 폴더 안 모든 csv 처리
#     python make_display_report.py  파일.csv  --cyc 5        → 수명시험 N사이클 간격 표시
#
#  파이썬 미설치 PC: portable python(무설치) + openpyxl 로 실행 가능.
# =====================================================================
import sys, os, csv, glob, datetime, argparse

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    sys.exit("openpyxl 가 필요합니다:  pip install openpyxl")

YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

# 데이터 표의 헤더(이 줄부터 아래가 측정 데이터)
DATA_HEADER = ["Step", "Status", "Time", "Step time", "Program time", "Cycle",
               "Cycle level", "Procedure", "Voltage", "Current",
               "AhStep", "AhCha", "AhDch", "AhAccu"]

# 시:분:초 문자열 → 엑셀 time / timedelta (경과시간은 24h 넘을 수 있어 timedelta)
def parse_time(s, elapsed=False):
    s = s.strip()
    if not s or ":" not in s:
        return s
    parts = s.split(":")
    try:
        h, m, sec = (int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0)
    except ValueError:
        return s
    if elapsed or h >= 24:
        return datetime.timedelta(hours=h, minutes=m, seconds=sec)
    return datetime.time(h, m, sec)

def to_num(s):
    s = s.strip()
    if s == "":
        return None
    try:
        if s.lstrip("-").replace(".", "", 1).isdigit():
            return float(s) if "." in s else int(s)
        return float(s)
    except ValueError:
        return s

# CSV 읽어 (메타행들, 데이터 헤더 행 index, 데이터행 리스트) 반환
def read_csv(path):
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        rows = list(csv.reader(f))
    hdr_idx = None
    for i, r in enumerate(rows):
        if r and r[0].strip() == "Step" and len(r) > 1 and r[1].strip() == "Status":
            hdr_idx = i
            break
    if hdr_idx is None:
        raise ValueError("데이터 헤더(Step,Status...)를 찾지 못함: " + path)
    meta = rows[:hdr_idx]                    # 헤더 블록(제목/배터리ID/프로그램 등)
    unit_row = rows[hdr_idx + 1]             # [V],[A] 단위 행
    data = rows[hdr_idx + 2:]                # 실제 측정 데이터
    # 뒤쪽 완전 빈 행 제거
    while data and not any(c.strip() for c in data[-1]):
        data.pop()
    return meta, rows[hdr_idx], unit_row, data

# 컬럼 인덱스 헬퍼
def col(header_row, name):
    for i, c in enumerate(header_row):
        if c.strip() == name:
            return i
    return -1

# ── 핵심: 표시할 행 결정 ─────────────────────────────────────────────
def pick_highlights(data, hdr, cyc_interval):
    """data: 원본 데이터행. 반환: 표시할 data index 집합."""
    ci_step   = col(hdr, "Step")
    ci_status = col(hdr, "Status")
    ci_cyc    = col(hdr, "Cycle")

    def g(r, i):
        return r[i].strip() if 0 <= i < len(r) else ""

    # 사이클 수로 시험 종류 판별
    max_cyc = 0
    for r in data:
        v = g(r, ci_cyc)
        if v.isdigit():
            max_cyc = max(max_cyc, int(v))

    hl = set()

    if max_cyc >= 2:
        # ── 수명/반복 시험: N 사이클 간격의 '방전 종료' 행 표시 ──
        # 각 사이클에서 마지막 DCH 행(그 사이클 방전용량 확정점)을 표시
        last_dch_of_cyc = {}
        for i, r in enumerate(data):
            if g(r, ci_status) == "DCH":
                cv = g(r, ci_cyc)
                if cv.isdigit():
                    last_dch_of_cyc[int(cv)] = i
        for c, i in last_dch_of_cyc.items():
            if c > 0 and c % cyc_interval == 0:
                hl.add(i)
        return hl, ("cycle", max_cyc)

    # ── 단일 사이클(용량/RC/충전수입성 등): 스텝 전환점 표시 ──
    # 스텝 번호로 블록 분할
    blocks = []
    for i, r in enumerate(data):
        st = g(r, ci_step)
        sta = g(r, ci_status)
        if blocks and blocks[-1][0] == st:
            blocks[-1][3] = i          # end
        else:
            blocks.append([st, sta, i, i])   # [step, status, start, end]

    # 앞쪽 휴지(PAU) 블록 제거(시험 전 대기), 뒤쪽 정지(STO)/9999 제거
    while blocks and blocks[0][1] == "PAU":
        blocks.pop(0)
    while blocks and (blocks[-1][1] == "STO" or blocks[-1][0] == "9999"):
        blocks.pop()

    if blocks:
        hl.add(blocks[0][2])           # 첫 시험 블록 시작(초기값)
        for b in blocks:
            hl.add(b[3])               # 각 블록 끝(충전/방전/휴지 완료점)
    return hl, ("single", max_cyc)

# ── 엑셀 쓰기 ────────────────────────────────────────────────────────
def write_xlsx(meta, hdr, unit_row, data, highlights, out_path, sheet_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # 시:분:초로 해석할 컬럼
    ci_time = col(hdr, "Time")
    ci_step_t = col(hdr, "Step time")
    ci_prog_t = col(hdr, "Program time")
    numeric_names = ["Step", "Cycle", "Cycle level", "Voltage", "Current",
                     "AhStep", "AhCha", "AhDch", "AhAccu"]
    numeric_cols = {col(hdr, n) for n in numeric_names if col(hdr, n) >= 0}

    r = 1
    # 메타 블록 그대로
    for row in meta:
        for c, val in enumerate(row):
            if val.strip():
                ws.cell(r, c + 1, val)
        r += 1
    # 데이터 헤더 + 단위행 (굵게)
    for c, val in enumerate(hdr):
        cell = ws.cell(r, c + 1, val)
        cell.font = Font(bold=True)
    r += 1
    for c, val in enumerate(unit_row):
        if val.strip():
            ws.cell(r, c + 1, val)
    r += 1

    data_start_row = r
    for di, row in enumerate(data):
        for c, val in enumerate(row):
            v = val.strip()
            if v == "":
                continue
            if c == ci_time:
                out = parse_time(v)             # 하루 안(시각)
            elif c in (ci_step_t, ci_prog_t):
                out = parse_time(v, elapsed=True)  # 경과(24h+)
            elif c in numeric_cols:
                out = to_num(v)
            else:
                out = v
            ws.cell(r, c + 1, out)
        if di in highlights:
            for c in range(len(hdr)):
                ws.cell(r, c + 1).fill = YELLOW
        r += 1

    # 컬럼 폭 살짝
    for c in range(1, len(hdr) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 11

    wb.save(out_path)
    return data_start_row

# ── 파일 1개 처리 ────────────────────────────────────────────────────
def process(path, cyc_interval, outdir=None):
    meta, hdr, unit_row, data = read_csv(path)
    highlights, (kind, maxc) = pick_highlights(data, hdr, cyc_interval)

    base = os.path.splitext(os.path.basename(path))[0]
    # 시트명 = 배터리ID 또는 파일명
    sheet = base
    if meta and meta[0]:
        title = meta[0][0]
        if "Batt" in title:
            sheet = base
    out_name = base + " - 표시.xlsx"
    dest_dir = outdir or os.path.dirname(path) or "."
    if dest_dir and not os.path.isdir(dest_dir):
        os.makedirs(dest_dir, exist_ok=True)
    out_path = os.path.join(dest_dir, out_name)
    write_xlsx(meta, hdr, unit_row, data, highlights, out_path, sheet)

    tag = "수명(%d사이클,%d칸 표시)" % (maxc, len(highlights)) if kind == "cycle" \
          else "단일시험(%d칸 표시)" % len(highlights)
    print("  OK  %-28s → %s   [%s]" % (os.path.basename(path), out_name, tag))
    return out_path

def main():
    ap = argparse.ArgumentParser(description="BTS-600 CSV → 측정값 표시 엑셀")
    ap.add_argument("target", help="CSV 파일 또는 폴더")
    ap.add_argument("--cyc", type=int, default=5,
                    help="수명시험: N사이클 간격으로 방전종료 표시 (기본 5)")
    ap.add_argument("--out", default=None, help="출력 폴더(기본: 원본과 같은 폴더)")
    args = ap.parse_args()

    if os.path.isdir(args.target):
        files = sorted(glob.glob(os.path.join(args.target, "*.csv")))
        # J2801 은 대상에서 제외
        skipped = [f for f in files if "j2801" in os.path.basename(f).lower()]
        files = [f for f in files if "j2801" not in os.path.basename(f).lower()]
        if not files:
            sys.exit("폴더에 처리할 CSV가 없습니다: " + args.target)
        print("%d개 CSV 처리 (J2801 제외 %d개):" % (len(files), len(skipped)))
        for f in files:
            try:
                process(f, args.cyc, args.out)
            except Exception as e:
                print("  ERR %s: %s" % (os.path.basename(f), e))
        for f in skipped:
            print("  SKIP %s (J2801 제외)" % os.path.basename(f))
    else:
        if "j2801" in os.path.basename(args.target).lower():
            print("  SKIP %s (J2801 제외)" % os.path.basename(args.target))
        else:
            process(args.target, args.cyc, args.out)
    print("완료.")

if __name__ == "__main__":
    main()
